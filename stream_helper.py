from itertools import count
from tkinter import *
from tkinter import ttk, filedialog
import pandas as pd
from openpyxl import Workbook
import datetime
#dummy data
dummy_data = [
	["matt820223", "小洋裝", "1", "藍色"],
	["allen9527", "小洋裝", "1", "黃色"],
	["cutecctwo", "小洋裝", "1", "藍色"],
	["tailspurple", "小洋裝", "1", "藍色"],
	["afeiori", "小洋裝", "1", "黃色"],
	["matt820223", "長褲", "1", "綠色"],
	["matt820223", "裙子", "1", "藍色"],
	["allen9527", "長褲", "1", "藍色"],
	["tailspurple", "長褲", "1", "藍色"],
	["cutecctwo", "長褲", "1", "藍色"],
	["allen9527", "裙子", "1", "藍色"],
	["afeiori", "裙子", "1", "藍色"],
	["dorisoa", "長褲", "1", "藍色"],
]





root = Tk()
root.title('Excel To Treeview')

root.geometry("1200x900")

#Create treeview frame
tree_frame = Frame(root)
tree_frame.pack(pady=20)

#Treeview scrollbar
tree_scroll = Scrollbar(tree_frame)
tree_scroll.pack(side=RIGHT, fill=Y)

#Create treeview
my_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set)
my_tree.pack(pady=20)

style = ttk.Style()
style.configure("Treeview", font=("Arial", 18), rowheight=35)
style.configure("Treeview.Heading", font=("Arial", 15))

#Configure the scrollbar
tree_scroll.config(command=my_tree.yview)

#Define columns
my_tree['columns'] = ("買家ID", "商品名稱", "數量", "顏色", "備註")
#Format columns
my_tree.column("#0", width=0, stretch=NO)
my_tree.column("買家ID", anchor=W, width=300)
my_tree.column("商品名稱",anchor=W, width=300)
my_tree.column("數量",anchor=W, width=100)
my_tree.column("顏色",anchor=W, width=100)
my_tree.column("備註",anchor=W, width=300)

my_tree.heading("#0", text="", anchor=W)
my_tree.heading("買家ID", text="買家ID", anchor=CENTER)
my_tree.heading("商品名稱", text="商品名稱", anchor=CENTER)
my_tree.heading("數量", text="數量", anchor=W)
my_tree.heading("顏色", text="顏色", anchor=W)
my_tree.heading("備註", text="備註", anchor=W)
#Add Data
count = 0
for data in dummy_data:
	my_tree.insert(parent='', index='end', iid=count, text="", values=(data[0], data[1], data[2], data[3], ""))
	count += 1



add_frame = Frame(root)
add_frame.pack(pady=20)

nl = Label(add_frame, text="買家ID", font=("Arial", 15))
nl.grid(row=0, column=0, padx=5)

il = Label(add_frame, text="商品名稱", font=("Arial", 15))
il.grid(row=0, column=1, padx=5)

tl = Label(add_frame, text="數量", font=("Arial", 15))
tl.grid(row=0, column=2, padx=5)

cl = Label(add_frame, text="顏色", font=("Arial", 15))
cl.grid(row=0, column=3, padx=5)

pl = Label(add_frame, text="備註", font=("Arial", 15))
pl.grid(row=0, column=4, padx=5)


cid_box = Entry(add_frame, font=("Arial", 15))
cid_box.grid(row=1, column=0, padx=5)

iname_box = Entry(add_frame, font=("Arial", 15))
iname_box.grid(row=1, column=1, padx=5)

number_box = Entry(add_frame, font=("Arial", 15))
number_box.grid(row=1, column=2, padx=5)

color_box = Entry(add_frame, font=("Arial", 15))
color_box.grid(row=1, column=3, padx=5)

note_box = Entry(add_frame, font=("Arial", 15))
note_box.grid(row=1, column=4, padx=5)


#Add Record
def add_record():
	global count
	my_tree.insert(parent='', index='end', iid=count, text="Parent", values=(cid_box.get(), iname_box.get(), number_box.get(), color_box.get(), note_box.get()))
	count += 1
	#clear the boxes
	cid_box.delete(0, END)
	iname_box.delete(0, END)
	number_box.delete(0, END)
	color_box.delete(0, END)
	note_box.delete(0, END)

#Remove all records
def remove_all():
	for record in my_tree.get_children():
		my_tree.delete(record)

#Remove one selected
def remove_one():
	select_one = my_tree.selection()[0]
	my_tree.delete(select_one)

#Select Record 
def select_record():
	#clear entry boxes
	cid_box.delete(0, END)
	iname_box.delete(0, END)
	number_box.delete(0, END)
	color_box.delete(0, END)
	note_box.delete(0, END)
	
	#Grab record number
	selected = my_tree.focus()

	#Grab record values
	values = my_tree.item(selected, 'values')
	cid_box.insert(0, values[0])
	iname_box.insert(0, values[1])
	number_box.insert(0, values[2])
	color_box.insert(0, values[3])
	note_box.insert(0, values[4])


#Save updated record
def update_record():
	#Grab record number
	selected = my_tree.focus()
	my_tree.item(selected, text="", values=(cid_box.get(), iname_box.get(), number_box.get(), color_box.get(), note_box.get()))

	#clear entry boxes
	cid_box.delete(0, END)
	iname_box.delete(0, END)
	number_box.delete(0, END)
	color_box.delete(0, END)
	note_box.delete(0, END)


#export to excel file
def export_excel():
	data_list = []

	wb = Workbook()	
	ws = wb.active
	ws.append(["買家ID", "商品名稱", "數量", "顏色", "備註"])
	for child in my_tree.get_children():
		data_list.append(my_tree.item(child)["values"])
		ws.append(my_tree.item(child)["values"])
	file_name = datetime.datetime.now().strftime("%m-%d-%Y_%H%M")
	df = pd.DataFrame(data_list, columns =['買家ID', '商品名稱', '數量', '顏色', '備註'])

	all_buyers = df['買家ID'].value_counts().index
	df_buyer = pd.DataFrame(columns =['買家ID', '商品名稱', '數量', '顏色'])
	for buyer in all_buyers:
		tmp = df[df['買家ID']==buyer]
		df_buyer = df_buyer.append(tmp)
	list_buyer = df_buyer.values.tolist()

	

	ws_1 = wb.create_sheet("Summary")
	ws_2 = wb.create_sheet("AllCustomer")
	for buyer in list_buyer:
		ws_2.append(buyer)
	wb.save(file_name + ".xlsx")

#Button 
add_record = Button(root, text="Add Record", command=add_record)
add_record.pack(pady=10)

#Select Button
select_record = Button(root, text="Select record", command=select_record)
select_record.pack(pady=10)
#update Button
update_record = Button(root, text="Update record", command=update_record)
update_record.pack(pady=10)

#Remove One Button
remove_one = Button(root, text="Remove one Selected", fg='white', bg='red', command=remove_one)
remove_one.pack(pady=20)

#Remove all Button
remove_all = Button(root, text="Remove All Records", fg='white', bg='red', command=remove_all)
remove_all.pack(pady=10)

#Export to excel Button

export_excel = Button(root, text="Export to excel", fg='white', bg='black', command=export_excel)
export_excel.pack(pady=10)



# temp_label = Label(root, text="")
# temp_label.pack(pady=20)






root.mainloop()
